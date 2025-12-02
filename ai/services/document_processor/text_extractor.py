import logging
from pathlib import Path
from typing import Optional
import asyncio

logger = logging.getLogger(__name__)


class TextExtractor:
    """Service for extracting text from various document formats"""
    
    def __init__(self):
        self.supported_extensions = {".pdf", ".docx", ".txt", ".xlsx", ".xls"}
        # Excel processing limits (configurable for performance)
        self.max_sheets_per_file = 20  # Limit sheets to prevent memory issues
        self.max_rows_per_sheet = 10000  # Limit rows per sheet
    
    async def extract_text_async(self, file_path: str) -> str:
        """Extract text asynchronously to avoid blocking"""
        def sync_extract():
            return self._extract_text_sync(file_path)
        
        # Run in thread pool for CPU-intensive operations
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, sync_extract)
    
    def _extract_text_sync(self, file_path: str) -> str:
        """Synchronous text extraction with better error handling"""
        path_obj = Path(file_path)
        ext = path_obj.suffix.lower()
        
        try:
            if ext == ".pdf":
                return self._extract_pdf(file_path)
            elif ext == ".docx":
                return self._extract_docx(file_path)
            elif ext == ".txt":
                return self._extract_txt(file_path)
            elif ext in {".xlsx", ".xls"}:
                return self._extract_excel(file_path)
            else:
                raise ValueError(f"Unsupported file type: {ext}")
        except Exception as e:
            logger.error(f"Text extraction failed for {file_path}: {e}")
            raise
    
    def _extract_pdf(self, file_path: str) -> str:
        """Extract text from PDF"""
        try:
            import fitz
            doc = fitz.open(file_path)
            text = ""
            for page in doc:
                text += page.get_text()
            doc.close()
            return text
        except Exception as e:
            logger.error(f"PDF extraction failed for {file_path}: {e}")
            raise
    
    def _extract_docx(self, file_path: str) -> str:
        """Extract text from DOCX"""
        try:
            import docx
            doc = docx.Document(file_path)
            return "\n".join([para.text for para in doc.paragraphs])
        except Exception as e:
            logger.error(f"DOCX extraction failed for {file_path}: {e}")
            raise
    
    def _extract_txt(self, file_path: str) -> str:
        """Extract text from TXT with encoding detection"""
        try:
            encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
            for encoding in encodings:
                try:
                    with open(file_path, "r", encoding=encoding) as f:
                        return f.read()
                except UnicodeDecodeError:
                    continue
            
            # Fallback with error handling
            with open(file_path, "r", encoding='utf-8', errors='ignore') as f:
                return f.read()
        except Exception as e:
            logger.error(f"TXT extraction failed for {file_path}: {e}")
            raise
    
    def _extract_excel(self, file_path: str) -> str:
        """
        Extract text from Excel files (.xlsx and .xls).
        
        Handles:
        - Multiple sheets (with sheet names as context)
        - Table structure preservation
        - Headers detection
        - Empty cells gracefully
        - Large files with configurable limits
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Structured text representation of Excel content
        """
        path_obj = Path(file_path)
        ext = path_obj.suffix.lower()
        
        if ext == ".xlsx":
            return self._extract_xlsx(file_path)
        elif ext == ".xls":
            return self._extract_xls(file_path)
        else:
            raise ValueError(f"Unsupported Excel format: {ext}")
    
    def _extract_xlsx(self, file_path: str) -> str:
        """Extract text from .xlsx files using openpyxl"""
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            text_parts = []
            
            # Process each sheet (with limit to prevent memory issues)
            sheet_names = workbook.sheetnames[:self.max_sheets_per_file]
            
            if len(workbook.sheetnames) > self.max_sheets_per_file:
                logger.warning(
                    f"Excel file {file_path} has {len(workbook.sheetnames)} sheets, "
                    f"processing only first {self.max_sheets_per_file}"
                )
            
            for sheet_name in sheet_names:
                sheet = workbook[sheet_name]
                sheet_text = self._extract_sheet_content(sheet, sheet_name)
                if sheet_text.strip():  # Only add non-empty sheets
                    text_parts.append(sheet_text)
            
            workbook.close()
            
            if not text_parts:
                logger.warning(f"No extractable content found in {file_path}")
                return ""
            
            return "\n\n".join(text_parts)
            
        except Exception as e:
            error_msg = str(e).lower()
            if "password" in error_msg or "encrypted" in error_msg:
                raise ValueError(f"Excel file is password-protected: {file_path}")
            elif "corrupt" in error_msg or "invalid" in error_msg:
                raise ValueError(f"Excel file appears to be corrupted: {file_path}")
            else:
                logger.error(f"XLSX extraction failed for {file_path}: {e}")
                raise
    
    def _extract_xls(self, file_path: str) -> str:
        """Extract text from legacy .xls files using xlrd"""
        try:
            import xlrd
            
            workbook = xlrd.open_workbook(file_path, on_demand=True)
            text_parts = []
            
            # Process each sheet (with limit)
            sheet_count = min(len(workbook.sheet_names()), self.max_sheets_per_file)
            
            if len(workbook.sheet_names()) > self.max_sheets_per_file:
                logger.warning(
                    f"Excel file {file_path} has {len(workbook.sheet_names())} sheets, "
                    f"processing only first {self.max_sheets_per_file}"
                )
            
            for sheet_idx in range(sheet_count):
                sheet = workbook.sheet_by_index(sheet_idx)
                sheet_name = sheet.name
                sheet_text = self._extract_xls_sheet_content(sheet, sheet_name)
                if sheet_text.strip():  # Only add non-empty sheets
                    text_parts.append(sheet_text)
            
            workbook.release_resources()
            
            if not text_parts:
                logger.warning(f"No extractable content found in {file_path}")
                return ""
            
            return "\n\n".join(text_parts)
            
        except xlrd.biffh.XLRDError as e:
            error_msg = str(e).lower()
            if "password" in error_msg or "encrypted" in error_msg:
                raise ValueError(f"Excel file is password-protected: {file_path}")
            elif "corrupt" in error_msg:
                raise ValueError(f"Excel file appears to be corrupted: {file_path}")
            else:
                logger.error(f"XLS extraction failed for {file_path}: {e}")
                raise
        except Exception as e:
            logger.error(f"XLS extraction failed for {file_path}: {e}")
            raise
    
    def _extract_sheet_content(self, sheet, sheet_name: str) -> str:
        """
        Extract content from an openpyxl sheet.
        
        Formats output as structured text with headers and rows.
        """
        lines = [f"Sheet: {sheet_name}"]
        
        # Get used range
        if sheet.max_row == 0 or sheet.max_column == 0:
            return ""  # Empty sheet
        
        # Limit rows to prevent memory issues
        max_row = min(sheet.max_row, self.max_rows_per_sheet)
        
        if sheet.max_row > self.max_rows_per_sheet:
            logger.debug(
                f"Sheet '{sheet_name}' has {sheet.max_row} rows, "
                f"processing only first {self.max_rows_per_sheet}"
            )
        
        # Extract headers (first row)
        headers = []
        for col_idx in range(1, min(sheet.max_column + 1, 100)):  # Limit columns too
            cell_value = sheet.cell(row=1, column=col_idx).value
            if cell_value is not None:
                headers.append(str(cell_value).strip())
            else:
                headers.append("")
        
        # Remove trailing empty headers
        while headers and not headers[-1]:
            headers.pop()
        
        if headers:
            lines.append(f"Headers: {' | '.join(headers)}")
        
        # Extract data rows
        row_count = 0
        for row_idx in range(2, max_row + 1):  # Start from row 2 (skip header)
            row_values = []
            has_content = False
            
            for col_idx in range(1, len(headers) + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    # Convert to string, handling different types
                    if isinstance(cell_value, (int, float)):
                        cell_str = str(cell_value)
                    elif isinstance(cell_value, str):
                        cell_str = cell_value.strip()
                    else:
                        cell_str = str(cell_value)
                    row_values.append(cell_str)
                    if cell_str:
                        has_content = True
                else:
                    row_values.append("")
            
            # Only add rows with content
            if has_content:
                lines.append(f"Row {row_idx - 1}: {' | '.join(row_values)}")
                row_count += 1
        
        if row_count == 0:
            return ""  # Sheet has no data rows
        
        return "\n".join(lines)
    
    def _extract_xls_sheet_content(self, sheet, sheet_name: str) -> str:
        """
        Extract content from an xlrd sheet (legacy .xls format).
        
        Formats output as structured text with headers and rows.
        """
        lines = [f"Sheet: {sheet_name}"]
        
        if sheet.nrows == 0 or sheet.ncols == 0:
            return ""  # Empty sheet
        
        # Limit rows
        max_row = min(sheet.nrows, self.max_rows_per_sheet)
        
        if sheet.nrows > self.max_rows_per_sheet:
            logger.debug(
                f"Sheet '{sheet_name}' has {sheet.nrows} rows, "
                f"processing only first {self.max_rows_per_sheet}"
            )
        
        # Extract headers (first row)
        headers = []
        for col_idx in range(min(sheet.ncols, 100)):  # Limit columns
            cell_value = sheet.cell_value(0, col_idx)
            if cell_value:
                headers.append(str(cell_value).strip())
            else:
                headers.append("")
        
        # Remove trailing empty headers
        while headers and not headers[-1]:
            headers.pop()
        
        if headers:
            lines.append(f"Headers: {' | '.join(headers)}")
        
        # Extract data rows
        row_count = 0
        for row_idx in range(1, max_row):  # Start from row 1 (skip header)
            row_values = []
            has_content = False
            
            for col_idx in range(len(headers)):
                if col_idx < sheet.ncols:
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    if cell_value:
                        cell_str = str(cell_value).strip()
                        row_values.append(cell_str)
                        if cell_str:
                            has_content = True
                    else:
                        row_values.append("")
                else:
                    row_values.append("")
            
            # Only add rows with content
            if has_content:
                lines.append(f"Row {row_idx}: {' | '.join(row_values)}")
                row_count += 1
        
        if row_count == 0:
            return ""  # Sheet has no data rows
        
        return "\n".join(lines)
    
    def is_supported_file(self, file_path: str) -> bool:
        """Check if file type is supported"""
        return Path(file_path).suffix.lower() in self.supported_extensions
