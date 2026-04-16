"""
SpreadsheetExtractor — row-group-aware chunk producer for XLS/XLSX files.

The regular DocumentChunker targets prose: it splits text at sentence or
paragraph boundaries targeting a 300-token window.  When an Excel file is
converted to plain text, that approach destroys table structure — a single
boundary can fall in the middle of a row block, so the LLM sees only partial
data in each context window.

This extractor bypasses the prose chunker entirely.  It reads each sheet,
groups rows into fixed-size windows (default 20 rows per chunk), and repeats
the column headers in every chunk so any chunk can be understood in isolation.

Chunk text format
-----------------
[File: report.xlsx | Sheet: Sheet1 | Rows: 1-20]
Headers: Doc ID | Type | Date | Amount
Row 1: GUA04 | Delivery Receipt | Aug 23 2025 | Php 1,775,767.50
Row 2: TCO002 | Delivery Receipt | Sep 01 2025 | Php 354,632.00
...

Integration
-----------
Call ``is_spreadsheet(file_path)`` to detect XLS/XLSX files, then replace the
text-extractor + prose-chunker pipeline with
``SpreadsheetExtractor.extract_chunks(file_path)``.
"""

import logging
from pathlib import Path
from typing import List, Tuple

from ..models import DocumentChunk

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Module-level constants
# ---------------------------------------------------------------------------

# Data rows per chunk.  At ~80 chars/row this gives ~1 600 chars (~400 tokens),
# well within the 512-token embedding model limit with room for the header line.
ROWS_PER_CHUNK: int = 20

SPREADSHEET_EXTENSIONS = frozenset({".xlsx", ".xls"})


def is_spreadsheet(file_path: str) -> bool:
    """Return True when *file_path* is an Excel file handled by this extractor."""
    return Path(file_path).suffix.lower() in SPREADSHEET_EXTENSIONS


# ---------------------------------------------------------------------------
# SpreadsheetExtractor
# ---------------------------------------------------------------------------

class SpreadsheetExtractor:
    """Converts XLS/XLSX files directly into DocumentChunk objects."""

    def __init__(self, rows_per_chunk: int = ROWS_PER_CHUNK) -> None:
        self.rows_per_chunk = rows_per_chunk

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def extract_chunks(self, file_path: str) -> Tuple[List[DocumentChunk], str]:
        """
        Extract a spreadsheet into DocumentChunk objects.

        Returns
        -------
        (chunks, content_preview)
            *content_preview* is the first 500 characters of the first chunk,
            suitable for storing in the document metadata DB.

        Raises
        ------
        ValueError
            If the file cannot be opened or has an unsupported extension.
        """
        ext = Path(file_path).suffix.lower()
        if ext == ".xlsx":
            chunks = self._extract_xlsx_chunks(file_path)
        elif ext == ".xls":
            chunks = self._extract_xls_chunks(file_path)
        else:
            raise ValueError(f"SpreadsheetExtractor does not handle {ext!r}")

        # Back-fill total_chunks now that the final count is known.
        total = len(chunks)
        for chunk in chunks:
            chunk.total_chunks = total

        preview = chunks[0].text[:500] if chunks else ""
        logger.info(
            "SpreadsheetExtractor: %s → %d chunk(s) from %d row-group(s)",
            Path(file_path).name,
            total,
            total,
        )
        return chunks, preview

    # ------------------------------------------------------------------
    # Format-specific extraction
    # ------------------------------------------------------------------

    def _extract_xlsx_chunks(self, file_path: str) -> List[DocumentChunk]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for .xlsx files: pip install openpyxl"
            ) from exc

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError(f"Cannot open XLSX file '{file_path}': {exc}") from exc

        chunks: List[DocumentChunk] = []
        chunk_id = 0
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            headers, rows = self._read_xlsx_sheet(sheet)
            new = self._rows_to_chunks(headers, rows, sheet_name, file_path, chunk_id)
            chunks.extend(new)
            chunk_id += len(new)

        wb.close()
        return chunks

    def _extract_xls_chunks(self, file_path: str) -> List[DocumentChunk]:
        try:
            import xlrd
        except ImportError as exc:
            raise ImportError(
                "xlrd is required for .xls files: pip install xlrd"
            ) from exc

        try:
            wb = xlrd.open_workbook(file_path, on_demand=True)
        except Exception as exc:
            raise ValueError(f"Cannot open XLS file '{file_path}': {exc}") from exc

        chunks: List[DocumentChunk] = []
        chunk_id = 0
        for sheet_idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(sheet_idx)
            headers, rows = self._read_xls_sheet(sheet)
            new = self._rows_to_chunks(headers, rows, sheet.name, file_path, chunk_id)
            chunks.extend(new)
            chunk_id += len(new)

        wb.release_resources()
        return chunks

    # ------------------------------------------------------------------
    # Sheet readers — return (headers, list-of-row-value-lists)
    # ------------------------------------------------------------------

    @staticmethod
    def _read_xlsx_sheet(sheet) -> Tuple[List[str], List[List[str]]]:
        """Read an openpyxl sheet into (headers, rows)."""
        headers: List[str] = []
        rows: List[List[str]] = []
        first_data_row = True

        for row_tuple in sheet.iter_rows(values_only=True):
            cells = [str(v).strip() if v is not None else "" for v in row_tuple]
            # Trim trailing empty cells.
            while cells and not cells[-1]:
                cells.pop()
            if not any(cells):
                continue  # skip blank rows

            if first_data_row:
                headers = cells
                first_data_row = False
            else:
                rows.append(cells)

        return headers, rows

    @staticmethod
    def _read_xls_sheet(sheet) -> Tuple[List[str], List[List[str]]]:
        """Read an xlrd sheet into (headers, rows)."""
        headers: List[str] = []
        rows: List[List[str]] = []

        for row_idx in range(sheet.nrows):
            cells = [
                str(sheet.cell_value(row_idx, col)).strip()
                if sheet.cell_value(row_idx, col)
                else ""
                for col in range(sheet.ncols)
            ]
            while cells and not cells[-1]:
                cells.pop()
            if not any(cells):
                continue

            if row_idx == 0:
                headers = cells
            else:
                rows.append(cells)

        return headers, rows

    # ------------------------------------------------------------------
    # Core: convert (headers, rows) into DocumentChunk objects
    # ------------------------------------------------------------------

    def _rows_to_chunks(
        self,
        headers: List[str],
        rows: List[List[str]],
        sheet_name: str,
        file_path: str,
        chunk_id_offset: int,
    ) -> List[DocumentChunk]:
        if not rows:
            return []

        file_name = Path(file_path).name
        header_line = ("Headers: " + " | ".join(headers)) if headers else ""
        chunks: List[DocumentChunk] = []

        for group_start in range(0, len(rows), self.rows_per_chunk):
            group = rows[group_start: group_start + self.rows_per_chunk]
            row_start = group_start + 1       # 1-based display index
            row_end = group_start + len(group)

            meta_line = (
                f"[File: {file_name} | Sheet: {sheet_name} | "
                f"Rows: {row_start}-{row_end}]"
            )
            row_lines = [
                f"Row {row_start + i}: {' | '.join(r)}"
                for i, r in enumerate(group)
            ]

            parts = [meta_line]
            if header_line:
                parts.append(header_line)
            parts.extend(row_lines)
            text = "\n".join(parts)

            chunks.append(
                DocumentChunk(
                    text=text,
                    chunk_id=chunk_id_offset + len(chunks),
                    total_chunks=0,  # back-filled in extract_chunks()
                    file_path=file_path,
                    start_pos=0,
                    end_pos=len(text),
                )
            )

        return chunks
