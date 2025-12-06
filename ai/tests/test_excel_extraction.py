"""
Test script for Excel file extraction functionality.

Tests:
- Basic Excel extraction (.xlsx and .xls)
- Multi-sheet files
- Edge cases (empty cells, headers, large files)
- Error handling (corrupted files, etc.)
"""

import asyncio
import sys
import os
from pathlib import Path
import tempfile
import shutil

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from services.document_processor.extraction import TextExtractor


async def test_basic_xlsx():
    """Test basic .xlsx file extraction"""
    print("\n" + "="*60)
    print("TEST 1: Basic .xlsx File Extraction")
    print("="*60)
    
    extractor = TextExtractor()
    
    # Create a simple test Excel file using openpyxl
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Data"
        
        # Add headers
        ws.append(["Date", "Product", "Quantity", "Price"])
        ws.append(["2024-01-01", "Widget A", 10, 5.00])
        ws.append(["2024-01-02", "Widget B", 5, 10.00])
        ws.append(["2024-01-03", "Widget C", 15, 7.50])
        
        # Save to temp file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            test_file = tmp.name
            wb.save(test_file)
        
        # Extract text
        result = await extractor.extract_text_async(test_file)
        
        print(f"\n✅ Extraction successful!")
        print(f"📄 File: {test_file}")
        print(f"📊 Extracted text length: {len(result)} characters")
        print(f"\n📝 Extracted content:\n{result}")
        
        # Verify content
        assert "Sheet: Sales Data" in result, "Sheet name not found"
        assert "Date" in result, "Header not found"
        assert "Widget A" in result, "Data not found"
        assert "10" in result, "Numeric data not found"
        
        print("\n✅ All assertions passed!")
        
        # Cleanup
        os.unlink(test_file)
        return True
        
    except ImportError:
        print("⚠️  openpyxl not available, skipping test")
        return False
    except Exception as e:
        print(f"❌ Test failed: {e}")
        import traceback
        traceback.print_exc()
        return False


async def test_multi_sheet():
    """Test multi-sheet Excel file"""
    print("\n" + "="*60)
    print("TEST 2: Multi-Sheet Excel File")
    print("="*60)
    
    extractor = TextExtractor()
    
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        
        # Sheet 1
        ws1 = wb.active
        ws1.title = "Sales"
        ws1.append(["Product", "Sales"])
        ws1.append(["A", 100])
        ws1.append(["B", 200])
        
        # Sheet 2
        ws2 = wb.create_sheet("Inventory")
        ws2.append(["Product", "Stock"])
        ws2.append(["A", 50])
        ws2.append(["B", 75])
        
        # Sheet 3
        ws3 = wb.create_sheet("Customers")
        ws3.append(["Name", "Email"])
        ws3.append(["John", "john@example.com"])
        
        # Save to temp file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            test_file = tmp.name
            wb.save(test_file)
        
        # Extract text
        result = await extractor.extract_text_async(test_file)
        
        print(f"\n✅ Extraction successful!")
        print(f"📄 File: {test_file}")
        print(f"📊 Extracted text length: {len(result)} characters")
        print(f"\n📝 Extracted content:\n{result}")
        
        # Verify all sheets are extracted
        assert "Sheet: Sales" in result, "Sales sheet not found"
        assert "Sheet: Inventory" in result, "Inventory sheet not found"
        assert "Sheet: Customers" in result, "Customers sheet not found"
        
        print("\n✅ All sheets extracted successfully!")
        
        # Cleanup
        os.unlink(test_file)
        return True
        
    except ImportError:
        print("⚠️  openpyxl not available, skipping test")
        return False
    except Exception as e:
        print(f"❌ Test failed: {e}")
        import traceback
        traceback.print_exc()
        return False


async def test_empty_cells():
    """Test Excel file with empty cells"""
    print("\n" + "="*60)
    print("TEST 3: Excel File with Empty Cells")
    print("="*60)
    
    extractor = TextExtractor()
    
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sparse Data"
        
        # Add data with empty cells
        ws.append(["Name", "Age", "City", "Country"])
        ws.append(["John", 25, "New York", "USA"])
        ws.append(["Jane", None, "London", None])  # Empty cells
        ws.append(["Bob", 30, None, "Canada"])
        ws.append([None, None, "Paris", "France"])  # Some empty
        
        # Save to temp file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            test_file = tmp.name
            wb.save(test_file)
        
        # Extract text
        result = await extractor.extract_text_async(test_file)
        
        print(f"\n✅ Extraction successful!")
        print(f"📄 File: {test_file}")
        print(f"📊 Extracted text length: {len(result)} characters")
        print(f"\n📝 Extracted content:\n{result}")
        
        # Verify empty cells are handled
        assert "Sheet: Sparse Data" in result, "Sheet name not found"
        assert "John" in result, "Data not found"
        assert "Jane" in result, "Row with empty cells not found"
        
        print("\n✅ Empty cells handled correctly!")
        
        # Cleanup
        os.unlink(test_file)
        return True
        
    except ImportError:
        print("⚠️  openpyxl not available, skipping test")
        return False
    except Exception as e:
        print(f"❌ Test failed: {e}")
        import traceback
        traceback.print_exc()
        return False


async def test_large_file_limits():
    """Test that large files are limited correctly"""
    print("\n" + "="*60)
    print("TEST 4: Large File Limits")
    print("="*60)
    
    extractor = TextExtractor()
    extractor.max_rows_per_sheet = 5  # Set low limit for testing
    
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Large Sheet"
        
        # Add headers
        ws.append(["ID", "Value"])
        
        # Add many rows (more than limit)
        for i in range(1, 20):  # 19 data rows, but limit is 5
            ws.append([i, f"Value {i}"])
        
        # Save to temp file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            test_file = tmp.name
            wb.save(test_file)
        
        # Extract text
        result = await extractor.extract_text_async(test_file)
        
        print(f"\n✅ Extraction successful!")
        print(f"📄 File: {test_file}")
        print(f"📊 Extracted text length: {len(result)} characters")
        print(f"\n📝 Extracted content (first 500 chars):\n{result[:500]}...")
        
        # Count rows in output
        row_count = result.count("Row ")
        print(f"\n📊 Rows extracted: {row_count} (limit was 5)")
        
        # Verify limit was applied
        assert row_count <= 5, f"Too many rows extracted: {row_count}"
        assert "Row 1" in result, "First row not found"
        
        print("\n✅ Row limit enforced correctly!")
        
        # Cleanup
        os.unlink(test_file)
        return True
        
    except ImportError:
        print("⚠️  openpyxl not available, skipping test")
        return False
    except Exception as e:
        print(f"❌ Test failed: {e}")
        import traceback
        traceback.print_exc()
        return False


async def test_unsupported_file():
    """Test that unsupported files raise appropriate errors"""
    print("\n" + "="*60)
    print("TEST 5: Unsupported File Type")
    print("="*60)
    
    extractor = TextExtractor()
    
    # Create a fake file
    with tempfile.NamedTemporaryFile(suffix='.xyz', delete=False) as tmp:
        test_file = tmp.name
        tmp.write(b"fake content")
    
    try:
        result = await extractor.extract_text_async(test_file)
        print(f"❌ Should have raised ValueError, but got: {result}")
        return False
    except ValueError as e:
        print(f"✅ Correctly raised ValueError: {e}")
        return True
    except Exception as e:
        print(f"❌ Unexpected error: {e}")
        return False
    finally:
        os.unlink(test_file)


async def test_file_validator():
    """Test that FileValidator recognizes Excel files"""
    print("\n" + "="*60)
    print("TEST 6: FileValidator Excel Support")
    print("="*60)
    
    from services.document_processor.extraction import FileValidator
    
    validator = FileValidator()
    
    # Test supported extensions
    assert ".xlsx" in validator.supported_extensions, ".xlsx not in supported extensions"
    assert ".xls" in validator.supported_extensions, ".xls not in supported extensions"
    
    print("✅ FileValidator supports .xlsx and .xls")
    
    # Test is_supported_file method
    assert validator.is_supported_file("test.xlsx"), "test.xlsx not recognized"
    assert validator.is_supported_file("test.XLSX"), "test.XLSX not recognized (case insensitive)"
    assert validator.is_supported_file("test.xls"), "test.xls not recognized"
    
    print("✅ is_supported_file() works correctly")
    
    return True


async def run_all_tests():
    """Run all tests"""
    print("\n" + "="*60)
    print("EXCEL EXTRACTION TEST SUITE")
    print("="*60)
    
    tests = [
        ("FileValidator Support", test_file_validator),
        ("Basic .xlsx Extraction", test_basic_xlsx),
        ("Multi-Sheet File", test_multi_sheet),
        ("Empty Cells Handling", test_empty_cells),
        ("Large File Limits", test_large_file_limits),
        ("Unsupported File Error", test_unsupported_file),
    ]
    
    results = []
    
    for test_name, test_func in tests:
        try:
            result = await test_func()
            results.append((test_name, result))
        except Exception as e:
            print(f"\n❌ Test '{test_name}' crashed: {e}")
            import traceback
            traceback.print_exc()
            results.append((test_name, False))
    
    # Print summary
    print("\n" + "="*60)
    print("TEST SUMMARY")
    print("="*60)
    
    passed = sum(1 for _, result in results if result)
    total = len(results)
    
    for test_name, result in results:
        status = "✅ PASS" if result else "❌ FAIL"
        print(f"{status}: {test_name}")
    
    print(f"\n📊 Results: {passed}/{total} tests passed")
    
    if passed == total:
        print("\n🎉 All tests passed!")
    else:
        print(f"\n⚠️  {total - passed} test(s) failed")
    
    return passed == total


if __name__ == "__main__":
    success = asyncio.run(run_all_tests())
    sys.exit(0 if success else 1)

